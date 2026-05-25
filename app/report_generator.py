"""
Report Generator Module
Creates compliance analysis reports in Excel and CSV formats.
Generates strong matches, weak matches, and missing clauses reports.
"""

import logging
import os
from typing import List, Dict, Optional
from datetime import datetime
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ReportGenerator:
    """
    Generates compliance analysis reports.
    
    Attributes:
        output_dir: Directory for report files
        include_csv: Also generate CSV versions
    """
    
    def __init__(self, output_dir: str = "outputs", include_csv: bool = True):
        """
        Initialize Report Generator.
        
        Args:
            output_dir: Output directory for reports
            include_csv: Generate CSV versions (default True)
        """
        self.output_dir = output_dir
        self.include_csv = include_csv
        
        # Create output directory
        os.makedirs(output_dir, exist_ok=True)
        
        logger.info(f"Report Generator initialized (output: {output_dir})")
    
    def generate_keyword_reports(
        self,
        connections: List[Dict],
        missing_keywords: List[Dict],
        output_prefix: str = "analysis"
    ) -> Dict[str, str]:
        """
        Generate keyword-based CSV reports (REFACTORED for keyword-per-row output).
        
        Each row in the output CSV = one keyword
        
        Args:
            connections: List of connection dicts with keyword-per-row structure
                - Each dict: {filename, keyword, connection_type, similarity_score, matched_reference, llm_explanation}
            missing_keywords: List of missing keyword dicts
                - Each dict: {filename, keyword, status, notes}
            output_prefix: Prefix for output files
            
        Returns:
            Dictionary mapping report type to file path
        """
        logger.info("Generating keyword-based CSV reports...")
        
        results = {}
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Generate connections CSV (keyword-per-row)
        connections_csv = self._generate_connections_csv(
            connections,
            output_prefix,
            timestamp
        )
        results["connections_csv"] = connections_csv
        
        # Generate missing CSV (missing keywords)
        missing_csv = self._generate_missing_csv(
            missing_keywords,
            output_prefix,
            timestamp
        )
        results["missing_csv"] = missing_csv
        
        logger.info(f"Generated keyword reports: {list(results.keys())}")
        return results
    
    def _generate_connections_csv(
        self,
        connections: List[Dict],
        output_prefix: str,
        timestamp: str
    ) -> str:
        """
        Generate connections CSV with keyword-per-row format.
        
        Columns:
        - filename: Document filename
        - keyword: The DO-178 keyword
        - connection_type: strong / weak / missing
        - similarity_score: Float 0-1 (None if keyword-only)
        - matched_reference: Reference text snippet
        - llm_explanation: Generated explanation from LLM
        
        Args:
            connections: Connection records
            output_prefix: Output file prefix
            timestamp: Timestamp for filename
            
        Returns:
            Path to generated CSV file
        """
        csv_filename = f"{self.output_dir}/{output_prefix}_connections_{timestamp}.csv"
        
        try:
            # Define expected columns
            columns = [
                "filename",
                "keyword",
                "connection_type",
                "similarity_score",
                "matched_reference",
                "llm_explanation"
            ]
            
            # Handle empty connections list
            if not connections:
                # Create empty DataFrame with correct columns
                df = pd.DataFrame(columns=columns)
            else:
                # Convert to DataFrame
                df = pd.DataFrame(connections)
                
                # Select only available columns in correct order
                df = df[[col for col in columns if col in df.columns]]
                
                # Sort by filename and connection type (strong first)
                if "connection_type" in df.columns:
                    type_order = {"strong": 0, "weak": 1, "missing": 2}
                    df["type_order"] = df["connection_type"].map(type_order)
                    df = df.sort_values(by=["filename", "type_order", "keyword"])
                    df = df.drop(columns=["type_order"])
            
            # Save to CSV
            df.to_csv(csv_filename, index=False, encoding='utf-8')
            logger.info(f"Generated connections CSV: {csv_filename} ({len(df)} rows)")
            
            return csv_filename
            
        except Exception as e:
            logger.error(f"Error generating connections CSV: {str(e)}")
            raise
    
    def _generate_missing_csv(
        self,
        missing_keywords: List[Dict],
        output_prefix: str,
        timestamp: str
    ) -> str:
        """
        Generate missing keywords CSV.
        
        Columns:
        - filename: Document filename
        - keyword: Missing keyword
        - status: Always "missing"
        - notes: Optional notes
        
        Args:
            missing_keywords: Missing keyword records
            output_prefix: Output file prefix
            timestamp: Timestamp for filename
            
        Returns:
            Path to generated CSV file
        """
        csv_filename = f"{self.output_dir}/{output_prefix}_missing_connections_{timestamp}.csv"
        
        try:
            # Define expected columns
            columns = [
                "filename",
                "keyword",
                "status",
                "notes"
            ]
            
            # Handle empty missing_keywords list
            if not missing_keywords:
                # Create empty DataFrame with correct columns
                df = pd.DataFrame(columns=columns)
            else:
                # Convert to DataFrame
                df = pd.DataFrame(missing_keywords)
                
                # Select only available columns
                df = df[[col for col in columns if col in df.columns]]
                
                # Sort by filename and keyword
                if len(df) > 0:
                    df = df.sort_values(by=["filename", "keyword"])
            
            # Save to CSV
            df.to_csv(csv_filename, index=False, encoding='utf-8')
            logger.info(f"Generated missing CSV: {csv_filename} ({len(df)} rows)")
            
            return csv_filename
            
        except Exception as e:
            logger.error(f"Error generating missing CSV: {str(e)}")
            raise
    
    def generate_reports(
        self,
        strong_matches: List[Dict],
        weak_matches: List[Dict],
        missing_clauses: List[Dict],
        project_name: str = "DO-178-Analysis",
        excel_path: Optional[str] = None,
        strong_csv_path: Optional[str] = None,
        weak_csv_path: Optional[str] = None,
        missing_csv_path: Optional[str] = None,
        summary_path: Optional[str] = None
    ) -> Dict[str, str]:
        """
        Generate all report files with optional custom paths.
        
        Args:
            strong_matches: List of strong match dicts
            weak_matches: List of weak match dicts
            missing_clauses: List of missing clause dicts
            project_name: Name for output files (used if custom paths not provided)
            excel_path: Custom path for Excel report (overrides default)
            strong_csv_path: Custom path for strong matches CSV
            weak_csv_path: Custom path for weak matches CSV
            missing_csv_path: Custom path for missing clauses CSV
            summary_path: Custom path for summary JSON
            
        Returns:
            Dictionary mapping report type to file path
            
        Example:
            # Use default paths
            reports = generator.generate_reports(strong, weak, missing)
            
            # Use custom paths
            reports = generator.generate_reports(
                strong, weak, missing,
                excel_path="C:/Reports/report.xlsx",
                strong_csv_path="C:/Reports/strong.csv"
            )
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = f"{project_name}_{timestamp}"
        
        results = {}
        
        # Generate Excel workbook
        excel_file = self._generate_excel(
            strong_matches,
            weak_matches,
            missing_clauses,
            base_name,
            custom_path=excel_path
        )
        results["excel"] = excel_file
        
        # Generate CSV files
        csv_results = {}
        if self.include_csv:
            csv_results = self._generate_csvs(
                strong_matches,
                weak_matches,
                missing_clauses,
                base_name,
                strong_csv_path=strong_csv_path,
                weak_csv_path=weak_csv_path,
                missing_csv_path=missing_csv_path
            )
            results.update(csv_results)
        
        # Generate summary report (pass actual generated file paths)
        files_map = {"excel": excel_file}
        files_map.update(csv_results)

        summary_file = self._generate_summary(
            strong_matches,
            weak_matches,
            missing_clauses,
            base_name,
            files_map=files_map,
            custom_path=summary_path
        )
        results["summary"] = summary_file
        
        logger.info(f"Generated reports: {list(results.keys())}")
        return results
    
    def _generate_excel(
        self,
        strong_matches: List[Dict],
        weak_matches: List[Dict],
        missing_clauses: List[Dict],
        base_name: str,
        custom_path: Optional[str] = None
    ) -> str:
        """
        Generate Excel workbook with multiple sheets.
        
        Args:
            strong_matches: Strong match records
            weak_matches: Weak match records
            missing_clauses: Missing clause records
            base_name: Base name for file
            
        Returns:
            Path to generated Excel file
        """
        if custom_path:
            excel_path = custom_path
            os.makedirs(os.path.dirname(custom_path) or ".", exist_ok=True)
        else:
            excel_path = os.path.join(self.output_dir, f"{base_name}_report.xlsx")
        
        try:
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create sheets
            self._create_strong_matches_sheet(wb, strong_matches)
            self._create_weak_matches_sheet(wb, weak_matches)
            self._create_missing_clauses_sheet(wb, missing_clauses)
            self._create_summary_sheet(wb, strong_matches, weak_matches, missing_clauses)
            
            # Save
            wb.save(excel_path)
            logger.info(f"Generated Excel report: {excel_path}")
            return excel_path
        
        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            raise
    
    def _create_strong_matches_sheet(self, wb: Workbook, matches: List[Dict]):
        """
        Create strong matches worksheet.
        
        Columns:
        - project_file
        - project_section
        - page_number
        - matched_clause
        - reference_section
        - similarity_score
        - explanation
        """
        ws = wb.create_sheet("Strong Matches")
        
        headers = [
            "Project File",
            "Project Section",
            "Page Number",
            "Matched Clause",
            "Reference Section",
            "Similarity Score",
            "Explanation"
        ]
        
        # Add headers
        ws.append(headers)
        
        # Format header
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Add data
        for match in matches:
            row = [
                match.get("project_file", ""),
                match.get("project_section", ""),
                match.get("page_number", ""),
                match.get("matched_clause", "")[:100],  # Truncate
                match.get("reference_section", ""),
                f"{match.get('similarity_score', 0):.3f}",
                match.get("explanation", "")[:100]
            ]
            ws.append(row)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 40
        
        # Wrap text
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    def _create_weak_matches_sheet(self, wb: Workbook, matches: List[Dict]):
        """
        Create weak matches worksheet.
        
        Columns:
        - project_file
        - project_section
        - candidate_clause
        - similarity_score
        - explanation
        """
        ws = wb.create_sheet("Weak Matches")
        
        headers = [
            "Project File",
            "Project Section",
            "Candidate Clause",
            "Similarity Score",
            "Explanation"
        ]
        
        ws.append(headers)
        
        # Format header
        header_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Add data
        for match in matches:
            row = [
                match.get("project_file", ""),
                match.get("project_section", ""),
                match.get("candidate_clause", "")[:100],
                f"{match.get('similarity_score', 0):.3f}",
                match.get("explanation", "")[:100]
            ]
            ws.append(row)
        
        # Adjust widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    def _create_missing_clauses_sheet(self, wb: Workbook, clauses: List[Dict]):
        """
        Create missing clauses worksheet.
        
        Columns:
        - clause
        - section
        - status
        - notes
        """
        ws = wb.create_sheet("Missing Clauses")
        
        headers = ["Clause", "Section", "Status", "Notes"]
        ws.append(headers)
        
        # Format header
        header_fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Add data
        for clause in clauses:
            row = [
                clause.get("clause", ""),
                clause.get("section", ""),
                clause.get("status", "Missing"),
                clause.get("notes", "")
            ]
            ws.append(row)
        
        # Adjust widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    def _create_summary_sheet(
        self,
        wb: Workbook,
        strong_matches: List[Dict],
        weak_matches: List[Dict],
        missing_clauses: List[Dict]
    ):
        """Create summary statistics sheet."""
        ws = wb.create_sheet("Summary", 0)
        
        total_strong = len(strong_matches)
        total_weak = len(weak_matches)
        total_missing = len(missing_clauses)
        total_analyzed = total_strong + total_weak + total_missing
        
        compliance_percentage = (
            (total_strong / total_analyzed * 100) if total_analyzed > 0 else 0
        )
        
        # Title
        ws['A1'] = "DO-178 Compliance Analysis Summary"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Statistics
        stats = [
            ("Analysis Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("", ""),
            ("Strong Matches", total_strong),
            ("Weak Matches", total_weak),
            ("Missing Clauses", total_missing),
            ("Total Analyzed", total_analyzed),
            ("", ""),
            ("Compliance Coverage", f"{compliance_percentage:.1f}%")
        ]
        
        row_num = 3
        for label, value in stats:
            ws[f'A{row_num}'] = label
            ws[f'B{row_num}'] = value
            
            if label and value:
                ws[f'A{row_num}'].font = Font(bold=True)
            
            row_num += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _generate_csvs(
        self,
        strong_matches: List[Dict],
        weak_matches: List[Dict],
        missing_clauses: List[Dict],
        base_name: str,
        strong_csv_path: Optional[str] = None,
        weak_csv_path: Optional[str] = None,
        missing_csv_path: Optional[str] = None
    ) -> Dict[str, str]:
        """
        Generate CSV files for each report type.
        
        Returns:
            Dictionary mapping type to file path
        """
        results = {}
        
        try:
            # Strong matches CSV
            if strong_matches:
                df_strong = pd.DataFrame(strong_matches)
                if strong_csv_path:
                    strong_csv = strong_csv_path
                    os.makedirs(os.path.dirname(strong_csv_path) or ".", exist_ok=True)
                else:
                    strong_csv = os.path.join(
                        self.output_dir,
                        f"{base_name}_strong_matches.csv"
                    )
                df_strong.to_csv(strong_csv, index=False, encoding='utf-8')
                results["strong_csv"] = strong_csv
                logger.info(f"Generated CSV: {strong_csv}")
            
            # Weak matches CSV
            if weak_matches:
                df_weak = pd.DataFrame(weak_matches)
                if weak_csv_path:
                    weak_csv = weak_csv_path
                    os.makedirs(os.path.dirname(weak_csv_path) or ".", exist_ok=True)
                else:
                    weak_csv = os.path.join(
                        self.output_dir,
                        f"{base_name}_weak_matches.csv"
                    )
                df_weak.to_csv(weak_csv, index=False, encoding='utf-8')
                results["weak_csv"] = weak_csv
                logger.info(f"Generated CSV: {weak_csv}")
            
            # Missing clauses CSV
            if missing_clauses:
                df_missing = pd.DataFrame(missing_clauses)
                if missing_csv_path:
                    missing_csv = missing_csv_path
                    os.makedirs(os.path.dirname(missing_csv_path) or ".", exist_ok=True)
                else:
                    missing_csv = os.path.join(
                        self.output_dir,
                        f"{base_name}_missing_clauses.csv"
                    )
                df_missing.to_csv(missing_csv, index=False, encoding='utf-8')
                results["missing_csv"] = missing_csv
                logger.info(f"Generated CSV: {missing_csv}")
            
            return results
        
        except Exception as e:
            logger.error(f"Error generating CSV files: {str(e)}")
            raise
    
    def _generate_summary(
        self,
        strong_matches: List[Dict],
        weak_matches: List[Dict],
        missing_clauses: List[Dict],
        base_name: str,
        files_map: Optional[Dict[str, str]] = None,
        custom_path: Optional[str] = None
    ) -> str:
        """
        Generate JSON summary report.
        
        Returns:
            Path to summary file
        """
        if custom_path:
            summary_path = custom_path
            os.makedirs(os.path.dirname(custom_path) or ".", exist_ok=True)
        else:
            summary_path = os.path.join(
                self.output_dir,
                f"{base_name}_summary.json"
            )
        
        try:
            total_items = len(strong_matches) + len(weak_matches) + len(missing_clauses)
            compliance_coverage = (
                (len(strong_matches) / total_items * 100) if total_items > 0 else 0
            )

            # Prefer actual generated file paths from files_map when provided
            files_generated: Dict[str, Optional[str]] = {
                "excel": None,
                "strong_matches_csv": None,
                "weak_matches_csv": None,
                "missing_clauses_csv": None
            }

            if files_map:
                files_generated["excel"] = files_map.get("excel")
                files_generated["strong_matches_csv"] = files_map.get("strong_csv") or files_map.get("strong_csv_path")
                files_generated["weak_matches_csv"] = files_map.get("weak_csv") or files_map.get("weak_csv_path")
                files_generated["missing_clauses_csv"] = files_map.get("missing_csv") or files_map.get("missing_csv_path")
            else:
                # fallback to base_name-derived names
                files_generated = {
                    "excel": f"{base_name}_report.xlsx",
                    "strong_matches_csv": f"{base_name}_strong_matches.csv" if strong_matches else None,
                    "weak_matches_csv": f"{base_name}_weak_matches.csv" if weak_matches else None,
                    "missing_clauses_csv": f"{base_name}_missing_clauses.csv" if missing_clauses else None
                }

            summary = {
                "timestamp": datetime.now().isoformat(),
                "statistics": {
                    "strong_matches": len(strong_matches),
                    "weak_matches": len(weak_matches),
                    "missing_clauses": len(missing_clauses),
                    "total_items": total_items,
                    "compliance_coverage": compliance_coverage
                },
                "files_generated": files_generated
            }

            with open(summary_path, 'w', encoding='utf-8') as f:
                json.dump(summary, f, indent=2)

            logger.info(f"Generated summary: {summary_path}")
            return summary_path

        except Exception as e:
            logger.error(f"Error generating summary: {str(e)}")
            raise
